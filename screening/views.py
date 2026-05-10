import os
import openpyxl
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy, reverse
from django.core.files.storage import default_storage
from django.utils.text import slugify
from django.utils import timezone
from datetime import datetime
from django.http import HttpResponse
from django.contrib import messages

from .models import RecruiterRoom, ResumeSubmission, Profile, GlobalSettings
from .forms import ExtendedUserCreationForm, AlgorithmSettingsForm, RoomEditForm 
from .utils import extract_text_from_pdf
from .screening_logic import execute_industry_screening

# --- 1. Authentication & Registration ---

def landing_page(request):
    if request.user.is_authenticated:
        try:
            if request.user.profile.role == 'RECRUITER':
                return redirect('dashboard')
            return redirect('home_ats_checker')
        except Profile.DoesNotExist:
            return render(request, 'screening/index.html', {'error': 'Profile missing.'})
    return render(request, 'screening/index.html')

def register(request):
    initial_role = request.GET.get('role', 'CANDIDATE').upper()
    if request.method == 'POST':
        form = ExtendedUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save() 
            login(request, user)
            if user.profile.role == 'RECRUITER':
                return redirect('recruiter_setup')
            return redirect('profile_setup')
    else:
        form = ExtendedUserCreationForm(initial={'role': initial_role})
    return render(request, 'registration/signup_unified.html', {'form': form})

class CustomLoginView(LoginView):
    template_name = 'registration/login.html' 
    def get_success_url(self):
        user = self.request.user
        if hasattr(user, 'profile'):
            return reverse_lazy('dashboard' if user.profile.role == 'RECRUITER' else 'home_ats_checker')
        return reverse_lazy('landing')

# --- 2. Profile Onboarding ---

@login_required
def profile_setup(request):
    profile = request.user.profile
    if request.method == 'POST':
        profile.full_name = request.POST.get('full_name')
        profile.phone_number = request.POST.get('phone')
        profile.linkedin_url = request.POST.get('linkedin')
        profile.bio = request.POST.get('bio')
        if 'profile_pic' in request.FILES:
            profile.profile_pic = request.FILES['profile_pic']
        profile.is_profile_complete = True
        profile.save()
        messages.success(request, "Candidate Profile Initialized!")
        return redirect('home_ats_checker')
    return render(request, 'screening/profile_setup.html', {'profile': profile})

@login_required
def recruiter_profile_setup(request):
    profile = request.user.profile 
    if request.method == 'POST':
        if 'profile_pic' in request.FILES:
            profile.profile_pic = request.FILES['profile_pic']
            profile.save()
            messages.success(request, "Profile picture updated!")
        return redirect('recruiter_setup')
    return render(request, 'screening/recruiter_setup.html', {'profile': profile})

# --- 3. Candidate Experience ---

@login_required
def home_ats_checker(request):
    if request.user.profile.role == 'RECRUITER':
        return redirect('dashboard')
        
    if request.method == 'POST':
        room_code = request.POST.get('room_slug')
        try:
            # Manually fetch the room instead of using get_object_or_404
            room = RecruiterRoom.objects.get(slug=room_code)
            return redirect('room_detail', slug=room.slug)
        except RecruiterRoom.DoesNotExist:
            # Render your custom alignment page if the code is wrong
            return render(request, 'screening/room_not_found.html', {
                'attempted_code': room_code
            })
            
    return render(request, 'screening/home.html')

@login_required
def candidate_history(request):
    submissions = ResumeSubmission.objects.filter(candidate=request.user).order_by('-submitted_at')
    return render(request, 'screening/candidate_history.html', {'submissions': submissions})

@login_required
def room_detail(request, slug):
    """
    Handles Room applications. 
    Includes DEBUG prints to verify if PDF text is being fetched correctly.
    """
    room = get_object_or_404(RecruiterRoom, slug=slug)
    
    # 1. Recruiter View
    if request.user.profile.role == 'RECRUITER':
        submissions = room.submissions.all().order_by('-score')
        return render(request, 'screening/room_admin.html', {'room': room, 'submissions': submissions})

    # 2. Candidate View: Check for existing application
    submission = ResumeSubmission.objects.filter(room=room, candidate=request.user).first()
    
    if submission:
        return render(request, 'screening/success.html', {
            'room': room, 
            'submission': submission, 
            'score': submission.score, 
            'missing_skills': submission.missing_skills.split(", ") if submission.missing_skills else []
        })

    # 3. Handle New Application
    if request.method == 'POST' and request.FILES.get('resume'):
        uploaded_file = request.FILES['resume']
        
        # STEP 1: Create the record immediately
        new_sub = ResumeSubmission.objects.create(
            room=room, 
            candidate=request.user, 
            resume_file=uploaded_file, 
            score=0
        )
        
        try:
            # STEP 2: Process with AI & DEBUG PRINT
            raw_text = extract_text_from_pdf(new_sub.resume_file.path)
            
            # --- DEBUG START: Confirm fetching ---
            print("--- DEBUG: EXTRACTED RESUME TEXT START ---")
            if raw_text:
                print(raw_text[:500]) # Print first 500 characters to terminal
            else:
                print("FAILED: raw_text is empty or None")
            print("--- DEBUG: EXTRACTED RESUME TEXT END ---")
            # --- DEBUG END ---

            results = execute_industry_screening(raw_text, room.jd_text, weights=room.get_weightage())
            
            # STEP 3: Update and save results
            new_sub.score = results['score']
            new_sub.skills = ", ".join(results.get('matched', []))
            new_sub.missing_skills = ", ".join(results.get('missing', []))
            new_sub.save()
            
            return render(request, 'screening/success.html', {
                'room': room, 
                'submission': new_sub, 
                'score': results['score'], 
                'missing_skills': results.get('missing', [])
            })
            
        except Exception as e:
            # STEP 4: Fallback - Log error to terminal but don't crash for the user
            print(f"--- DEBUG: AI Analysis error: {e} ---")
            
            return render(request, 'screening/success.html', {
                'room': room, 
                'submission': new_sub, 
                'score': "Pending", 
                'missing_skills': [],
                'analysis_error': True 
            })
            
    # Default: Show the room application page
    return render(request, 'screening/room_detail.html', {'room': room})

# --- 4. Recruiter Dashboard & Management ---

@login_required
def dashboard(request):
    if request.user.profile.role == 'CANDIDATE':
        return redirect('home_ats_checker')
    rooms = RecruiterRoom.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'screening/dashboard.html', {'rooms': rooms, 'now': timezone.now()})

@login_required
def create_room(request):
    if request.user.profile.role == 'CANDIDATE':
        return redirect('home_ats_checker')
    success_created = False
    if request.method == 'POST':
        name = request.POST.get('name')
        expires_at_raw = request.POST.get('expires_at')
        expires_at = timezone.make_aware(datetime.strptime(expires_at_raw, '%Y-%m-%dT%H:%M')) if expires_at_raw else None
        RecruiterRoom.objects.create(
            created_by=request.user, name=name, slug=slugify(name),
            jd_text=request.POST.get('jd_text'), expires_at=expires_at
        )
        success_created = True 
    existing_slugs = list(RecruiterRoom.objects.values_list('slug', flat=True))
    return render(request, 'screening/create_room.html', {'success_created': success_created, 'existing_slugs_json': json.dumps(existing_slugs)})

@login_required
def edit_room(request, slug):
    room = get_object_or_404(RecruiterRoom, slug=slug, created_by=request.user)
    if request.method == 'POST':
        form = RoomEditForm(request.POST, instance=room)
        if form.is_valid():
            room = form.save(commit=False)
            weights = {'skill_weight': form.cleaned_data.get('skill_weight', 60), 'experience_weight': form.cleaned_data.get('exp_weight', 40)}
            room.custom_weightage = weights
            room.save()
            messages.success(request, f"'{room.name}' settings updated.")
            return redirect('dashboard')
    else:
        weights = room.custom_weightage or {'skill_weight': 60, 'experience_weight': 40}
        form = RoomEditForm(instance=room, initial={'skill_weight': weights.get('skill_weight'), 'exp_weight': weights.get('experience_weight')})
    return render(request, 'screening/edit_room.html', {'form': form, 'room': room})

@login_required
def bulk_status_update(request):
    if request.method == 'POST':
        ids = request.POST.getlist('selected_ids')
        new_status = request.POST.get('new_status')
        if ids and new_status:
            submissions = ResumeSubmission.objects.filter(id__in=ids)
            if new_status == 'INTERVIEW':
                submissions.update(status=new_status, interview_date=request.POST.get('interview_date'), interview_time=request.POST.get('interview_time'), interview_type=request.POST.get('interview_type'), interview_location=request.POST.get('interview_location'))
            else:
                submissions.update(status=new_status)
            messages.success(request, f"Updated {submissions.count()} candidates.")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

# --- 5. Analytics & Export ---

@login_required
def find_matches(request, submission_id):
    """Restore the missing find_matches view."""
    original_sub = get_object_or_404(ResumeSubmission, id=submission_id, candidate=request.user)
    resume_text = extract_text_from_pdf(original_sub.resume_file.path)
    other_rooms = RecruiterRoom.objects.filter(expires_at__gt=timezone.now()).exclude(id=original_sub.room.id)
    match_results = []
    for room in other_rooms:
        score_data = execute_industry_screening(resume_text, room.jd_text, weights=room.get_weightage())
        if score_data['score'] >= 50:
            match_results.append({'room': room, 'score': score_data['score'], 'matched_skills': score_data['matched'][:5]})
    return render(request, 'screening/suggested_rooms.html', {'original_room': original_sub.room, 'matches': sorted(match_results, key=lambda x: x['score'], reverse=True)})

@login_required
def compare_skills(request, submission_id):
    submission = get_object_or_404(ResumeSubmission, id=submission_id)
    return render(request, 'screening/compare_skills.html', {
        'submission': submission, 'room': submission.room,
        'matched_skills': sorted(submission.skills.split(", ")) if submission.skills else [],
        'missing_skills': sorted(submission.missing_skills.split(", ")) if submission.missing_skills else [],
        'ai_insight': "Strong Match" if submission.score >= 80 else "Potential" if submission.score >= 50 else "Weak"
    })

@login_required
def export_to_excel(request):
    """
    Final Optimized Export: Bridges the gap between 0% scores and visible missing skills.
    Ensures recruiters see exactly WHY a candidate failed to match.
    """
    if request.method == 'POST':
        ids = request.POST.getlist('selected_ids')
        if not ids:
            messages.warning(request, "Please select at least one candidate to export.")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
            
        # Optimization: Fetching everything in one go
        submissions = ResumeSubmission.objects.filter(id__in=ids).select_related(
            'candidate', 
            'candidate__profile',
            'room'
        ).order_by('-score')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidate Analysis"
        
        # 1. Header Styling
        headers = [
            'Rank', 'Candidate Name', 'Email', 'Phone', 
            'Match Score', 'Matched Skills', 'Missing Skills (Gaps)', 'Applied Date'
        ]
        ws.append(headers)
        
        header_fill = openpyxl.styles.PatternFill(start_color="F1F5F9", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # 2. Data Processing
        for i, sub in enumerate(submissions, 1):
            profile = getattr(sub.candidate, 'profile', None)
            
            # Matched Skills Formatting
            matched_list = sub.skills.strip() if (sub.skills and sub.skills.strip()) else "No matches"
            
            # Logic Correction: Ensure "Gaps" show up even if match is 0%
            # If sub.missing_skills is empty but score is low, it means the JD might be empty
            if not sub.missing_skills or sub.missing_skills.strip() == "":
                if sub.score < 10.0:
                    gaps_list = "Check Room JD for keywords (None found)"
                else:
                    gaps_list = "No gaps identified"
            else:
                gaps_list = sub.missing_skills.strip()

            ws.append([
                i,
                profile.full_name if profile and profile.full_name else sub.candidate.username,
                sub.candidate.email,
                profile.phone_number if profile and profile.phone_number else "N/A",
                f"{sub.score}%",
                matched_list,
                gaps_list,
                sub.submitted_at.strftime('%Y-%m-%d')
            ])

        # 3. Dynamic Column Auto-Sizing
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Set width with a bit of padding, cap at 60 for long skill lists
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        # 4. Generate Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="ResuMetric_Report_{timestamp}.xlsx"'
        
        wb.save(response)
        return response

    return redirect('dashboard')

@login_required
def delete_room(request, room_id):
    room = get_object_or_404(RecruiterRoom, id=room_id, created_by=request.user)
    room.delete()
    messages.success(request, "Room deleted.")
    return redirect('dashboard')

@login_required
def recruiter_settings_view(request):
    settings_obj = GlobalSettings.objects.filter(is_active=True).first()
    if request.method == 'POST':
        form = AlgorithmSettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Global AI settings updated.")
            return redirect('recruiter_settings_view')
    else:
        form = AlgorithmSettingsForm(instance=settings_obj)
    return render(request, 'screening/settings_dashboard.html', {'form': form})

@login_required
def unlock_premium_features(request, submission_id):
    submission = get_object_or_404(ResumeSubmission, id=submission_id, candidate=request.user)
    submission.is_premium_unlocked = True
    submission.save()
    return redirect('room_detail', slug=submission.room.slug)